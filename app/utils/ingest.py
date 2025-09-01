import os
import sys
import uuid
import logging
from pathlib import Path

from app.utils.chroma_utils import get_chroma_collection

from langchain_community.document_loaders import (
    PyPDFLoader, TextLoader, Docx2txtLoader
)
from langchain.text_splitter import RecursiveCharacterTextSplitter
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

collection = get_chroma_collection()

def load_document(file_path: str):
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return PyPDFLoader(file_path).load()
    elif ext == ".txt":
        return TextLoader(file_path, encoding="utf-8").load()
    elif ext == ".docx":
        return Docx2txtLoader(file_path).load()
    elif ext == ".xlsx":
        return load_excel(file_path)
    else:
        raise ValueError(f"❌ Unsupported file type: {ext}")

def load_excel(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(
                [str(cell) if cell is not None else "" for cell in row]
            )
            text_rows.append(row_text)
        sheet_text = "\n".join(text_rows)
        docs.append({
            "page_content": sheet_text,
            "metadata": {"source": file_path, "sheet": sheet_name},
        })
    return docs

def split_documents(documents, chunk_size=1000, chunk_overlap=100):
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size, chunk_overlap=chunk_overlap
    )
    if hasattr(documents[0], "page_content"):
        return splitter.split_documents(documents)
    
    class FakeDoc:
        def __init__(self, page_content, metadata):
            self.page_content = page_content
            self.metadata = metadata
            
    fake_docs = [FakeDoc(d["page_content"], d["metadata"]) for d in documents]
    return splitter.split_documents(fake_docs)

def ingest_file(file_path: str):
    logger.info(f"📂 Memuat dokumen: {file_path} ...")
    documents = load_document(file_path)
    chunks = split_documents(documents)
    ids = [str(uuid.uuid4()) for _ in range(len(chunks))]
    texts = [c.page_content for c in chunks]
    metadatas = [c.metadata for c in chunks]

    logger.info(f"🔎 Memasukkan {len(chunks)} potongan ke ChromaDB ...")
    collection.add(ids=ids, documents=texts, metadatas=metadatas)
    logger.info(f"✅ Selesai ingest: {file_path}")

if __name__ == "__main__":
    sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
    
    if len(sys.argv) < 2:
        print("Penggunaan: python app/utils/ingest.py <path_ke_file>")
        sys.exit(1)
        
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"❌ File tidak ditemukan: {file_path}")
        sys.exit(1)
        
    ingest_file(file_path)
